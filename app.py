import streamlit as st
import pandas as pd
import io
from utils import load_multiple_excels, find_auto_join_columns, smart_join_dfs, get_df_summary, find_common_cols, do_join
from graph import analyze_query_structured
import os
from dotenv import load_dotenv

load_dotenv()

# ── Must be first ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ExcelLens — Excel Intelligence",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Design System ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500&family=IBM+Plex+Sans:wght@300;400;500;600&family=Syne:wght@700;800&display=swap');

*, *::before, *::after { box-sizing: border-box; }

html, body, [data-testid="stAppViewContainer"] {
    background: #0a0e1a !important;
    color: #e2e8f0 !important;
    font-family: 'IBM Plex Sans', sans-serif !important;
}
[data-testid="stAppViewContainer"] > .main { background: #0a0e1a !important; }
[data-testid="stSidebar"] {
    background: #0d1117 !important;
    border-right: 1px solid #1e2d3d !important;
}
[data-testid="stSidebar"] > div { padding: 0 !important; }
#MainMenu, footer, header { visibility: hidden; }
.stDeployButton { display: none; }

::-webkit-scrollbar { width: 4px; height: 4px; }
::-webkit-scrollbar-track { background: #0d1117; }
::-webkit-scrollbar-thumb { background: #2563eb; border-radius: 2px; }

.wordmark-bar {
    display: flex; align-items: center; gap: 12px;
    padding: 28px 0 20px 0;
    border-bottom: 1px solid #1e2d3d;
    margin-bottom: 28px;
}
.wordmark-icon {
    width: 36px; height: 36px; background: #2563eb;
    border-radius: 8px; display: flex; align-items: center;
    justify-content: center; font-size: 18px; flex-shrink: 0;
}
.wordmark-text { font-family: 'Syne', sans-serif; font-size: 1.25rem; font-weight: 800; color: #f8fafc; letter-spacing: -0.02em; line-height: 1; }
.wordmark-sub  { font-family: 'IBM Plex Mono', monospace; font-size: 0.6rem; color: #475569; letter-spacing: 0.12em; text-transform: uppercase; margin-top: 3px; }

.sidebar-label {
    font-family: 'IBM Plex Mono', monospace; font-size: 0.6rem;
    font-weight: 500; letter-spacing: 0.15em; text-transform: uppercase;
    color: #475569; margin: 24px 0 10px 0; padding-left: 2px;
}

.status-pill {
    display: inline-flex; align-items: center; gap: 6px;
    background: #0f2942; border: 1px solid #1e4080; color: #60a5fa;
    padding: 5px 12px; border-radius: 100px;
    font-family: 'IBM Plex Mono', monospace; font-size: 0.7rem; font-weight: 500; margin-bottom: 8px;
}
.status-pill .dot {
    width: 6px; height: 6px; background: #22c55e; border-radius: 50%;
    box-shadow: 0 0 6px #22c55e; animation: pulse 2s infinite;
}
@keyframes pulse { 0%, 100% { opacity: 1; } 50% { opacity: 0.4; } }

.section-eyebrow { font-family: 'IBM Plex Mono', monospace; font-size: 0.65rem; letter-spacing: 0.2em; text-transform: uppercase; color: #2563eb; margin-bottom: 6px; }
.section-title   { font-family: 'Syne', sans-serif; font-size: 1.6rem; font-weight: 800; color: #f8fafc; letter-spacing: -0.03em; line-height: 1.1; margin-bottom: 20px; }

.metric-row { display: flex; gap: 12px; margin-bottom: 24px; }
.metric-card { flex: 1; background: #0d1117; border: 1px solid #1e2d3d; border-radius: 10px; padding: 18px 20px; position: relative; overflow: hidden; }
.metric-card::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 2px; background: #2563eb; }
.metric-label { font-family: 'IBM Plex Mono', monospace; font-size: 0.62rem; letter-spacing: 0.14em; text-transform: uppercase; color: #475569; margin-bottom: 8px; }
.metric-value { font-family: 'Syne', sans-serif; font-size: 1.8rem; font-weight: 800; color: #f8fafc; letter-spacing: -0.03em; line-height: 1; }
.metric-card.accent::before { background: #7c3aed; }
.metric-card.green::before  { background: #059669; }
.metric-card.amber::before  { background: #d97706; }

.join-badge {
    display: inline-flex; align-items: center; gap: 5px;
    background: #1a1f35; border: 1px solid #2d3a5c; border-radius: 6px;
    padding: 3px 9px; font-family: 'IBM Plex Mono', monospace;
    font-size: 0.65rem; color: #818cf8; margin: 3px 3px 3px 0;
}

/* ── Answer summary box ── */
.answer-summary {
    background: #080c14; border-left: 3px solid #2563eb;
    border-radius: 0 8px 8px 0; padding: 10px 16px; margin-bottom: 12px;
    font-family: 'IBM Plex Sans', sans-serif; font-size: 0.88rem;
    color: #bfdbfe; line-height: 1.5;
}

/* ── Section divider inside message ── */
.msg-divider {
    border: none; border-top: 1px solid #1e2d3d;
    margin: 12px 0;
}

/* ── Follow-up label ── */
.followup-label {
    font-family: 'IBM Plex Mono', monospace; font-size: 0.58rem;
    letter-spacing: 0.14em; text-transform: uppercase;
    color: #334155; margin-bottom: 6px;
}

/* ── Export label ── */
.export-label {
    font-family: 'IBM Plex Mono', monospace; font-size: 0.58rem;
    letter-spacing: 0.14em; text-transform: uppercase; color: #334155;
}

/* Download button — ghost */
[data-testid="stDownloadButton"] > button {
    background: #0d1117 !important; color: #60a5fa !important;
    border: 1px solid #1e3a5f !important; border-radius: 6px !important;
    font-family: 'IBM Plex Mono', monospace !important; font-size: 0.68rem !important;
    font-weight: 500 !important; padding: 4px 14px !important;
    letter-spacing: 0.03em !important; transition: background 0.15s, border-color 0.15s !important;
}
[data-testid="stDownloadButton"] > button:hover {
    background: #0f1a2e !important; border-color: #2563eb !important; transform: none !important;
}

/* Follow-up buttons — slim ghost */
.followup-btn [data-testid="stButton"] > button,
.followup-btn .stButton > button {
    background: #0a0f1f !important; color: #818cf8 !important;
    border: 1px solid #2d3a5c !important; border-radius: 20px !important;
    font-family: 'IBM Plex Sans', sans-serif !important; font-size: 0.75rem !important;
    font-weight: 400 !important; padding: 5px 14px !important;
    white-space: normal !important; text-align: left !important;
}
.followup-btn .stButton > button:hover {
    background: #0f1a35 !important; border-color: #818cf8 !important; transform: none !important;
}

.stSelectbox > div > div { background: #0d1117 !important; border: 1px solid #1e2d3d !important; border-radius: 8px !important; color: #e2e8f0 !important; font-family: 'IBM Plex Sans', sans-serif !important; }
.stSelectbox > div > div:hover { border-color: #2563eb !important; }

.stButton > button {
    background: #2563eb !important; color: #fff !important; border: none !important;
    border-radius: 8px !important; font-family: 'IBM Plex Sans', sans-serif !important;
    font-weight: 600 !important; font-size: 0.82rem !important; letter-spacing: 0.01em !important;
    padding: 10px 18px !important; transition: background 0.15s, transform 0.1s !important;
}
.stButton > button:hover { background: #1d4ed8 !important; transform: translateY(-1px) !important; }
.stButton > button:active { transform: translateY(0) !important; }

[data-testid="stFileUploader"] { background: #0d1117 !important; border: 1px dashed #1e2d3d !important; border-radius: 10px !important; padding: 12px !important; }
[data-testid="stFileUploader"]:hover { border-color: #2563eb !important; }
[data-testid="stFileUploader"] label { color: #94a3b8 !important; font-family: 'IBM Plex Sans', sans-serif !important; font-size: 0.82rem !important; }

[data-testid="stDataFrame"] { border: 1px solid #1e2d3d !important; border-radius: 10px !important; overflow: hidden !important; }
.stDataFrame { background: #0d1117 !important; }

[data-testid="stExpander"] { background: #0d1117 !important; border: 1px solid #1e2d3d !important; border-radius: 10px !important; }
[data-testid="stExpander"] summary { font-family: 'IBM Plex Sans', sans-serif !important; font-weight: 600 !important; font-size: 0.85rem !important; color: #94a3b8 !important; }
[data-testid="stExpander"] summary:hover { color: #e2e8f0 !important; }

hr { border-color: #1e2d3d !important; margin: 32px 0 !important; }

[data-testid="stChatMessage"] { background: transparent !important; padding: 14px 20px !important; border-bottom: 1px solid #111827 !important; }
[data-testid="stChatMessage"]:last-child { border-bottom: none !important; }

[data-testid="stChatInputContainer"] { background: #0d1117 !important; border: 1px solid #1e2d3d !important; border-radius: 10px !important; margin-top: 10px !important; }
[data-testid="stChatInputContainer"]:focus-within { border-color: #2563eb !important; box-shadow: 0 0 0 3px rgba(37,99,235,0.15) !important; }
[data-testid="stChatInput"] { background: transparent !important; color: #e2e8f0 !important; font-family: 'IBM Plex Sans', sans-serif !important; font-size: 0.88rem !important; }

.user-msg { background: #1e3a5f; border: 1px solid #2563eb; border-radius: 12px 12px 4px 12px; padding: 10px 16px; color: #bfdbfe; font-size: 0.88rem; line-height: 1.5; display: inline-block; max-width: 100%; }

[data-testid="stChatMessage"][data-testid*="assistant"] { background: #080c14 !important; }
[data-testid="stAlert"] { background: #0f2942 !important; border: 1px solid #1e4080 !important; border-radius: 8px !important; color: #93c5fd !important; }

.stCaption { font-family: 'IBM Plex Mono', monospace !important; font-size: 0.68rem !important; color: #475569 !important; }

.nav-item { display: flex; align-items: center; gap: 10px; padding: 9px 14px; border-radius: 8px; margin-bottom: 4px; font-size: 0.82rem; font-weight: 500; color: #64748b; cursor: default; transition: background 0.15s; }
.nav-item.active { background: #0f2133; color: #60a5fa; border: 1px solid #1e4080; }
.nav-item .nav-icon { font-size: 1rem; width: 20px; text-align: center; }

.empty-state { text-align: center; padding: 60px 20px; color: #334155; }
.empty-state .empty-icon { font-size: 3rem; margin-bottom: 16px; }
.empty-state .empty-title { font-family: 'Syne', sans-serif; font-size: 1.1rem; font-weight: 700; color: #475569; margin-bottom: 8px; }
.empty-state .empty-sub { font-size: 0.8rem; color: #334155; font-family: 'IBM Plex Mono', monospace; }

.stSpinner > div { color: #2563eb !important; }
.element-container .stSuccess { background: #052e16 !important; border-color: #166534 !important; color: #86efac !important; }

.col-tag { display: inline-block; background: #111827; border: 1px solid #1e2d3d; border-radius: 4px; padding: 2px 8px; font-family: 'IBM Plex Mono', monospace; font-size: 0.65rem; color: #818cf8; margin: 2px; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(columns: list, rows: list, summary: str, question: str) -> bytes:
    """
    Build a styled .xlsx from real columns + rows returned by the structured agent.
    Sheet 1 "Results" — the actual data table
    Sheet 2 "Info"    — question + summary
    """
    output = io.BytesIO()

    # Build DataFrame from dynamic columns and rows
    try:
        df_result = pd.DataFrame(rows, columns=columns)
    except Exception:
        df_result = pd.DataFrame({"Answer": [str(r) for r in rows]})

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_result.to_excel(writer, sheet_name="Results", index=False)

        df_info = pd.DataFrame({"Field": ["Question", "Summary"], "Value": [question, summary]})
        df_info.to_excel(writer, sheet_name="Info", index=False)

        # Styling
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        wrap        = Alignment(wrap_text=True, vertical="top")

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = wrap
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = wrap
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)

    return output.getvalue()


def render_answer(chat: dict, idx: int, is_latest: bool):
    """Render a single assistant chat entry: summary + table + export + follow-ups."""
    structured = chat.get("structured", {})
    summary    = structured.get("summary", "")
    columns    = structured.get("columns", [])
    rows       = structured.get("rows", [])
    followups  = structured.get("followups", [])
    question   = chat.get("question", "")
    error      = structured.get("error")

    if error:
        st.error(f"Analysis error: {error}")
        return

    # ── Summary sentence ──────────────────────────────────────────────────────
    if summary:
        st.markdown(f'<div class="answer-summary">{summary}</div>', unsafe_allow_html=True)

    # ── Data table ────────────────────────────────────────────────────────────
    if columns and rows:
        try:
            df_result = pd.DataFrame(rows, columns=columns)
            st.dataframe(df_result, width='stretch', hide_index=True)
        except Exception as e:
            st.warning(f"Could not render table: {e}")
            st.write(rows)
    elif not summary:
        st.info("No data returned.")

    if not is_latest:
        return  # Export + follow-ups only on the most recent answer

    # ── Export ────────────────────────────────────────────────────────────────
    if columns and rows:
        st.markdown('<hr class="msg-divider">', unsafe_allow_html=True)
        ec1, ec2 = st.columns([1, 5])
        with ec1:
            st.markdown('<div class="export-label">Export</div>', unsafe_allow_html=True)
        with ec2:
            excel_bytes = build_excel(columns, rows, summary, question)
            st.download_button(
                label="⬇ Download Excel",
                data=excel_bytes,
                file_name="ExcelLens_result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_{idx}",
            )

    # ── Follow-up suggestions ─────────────────────────────────────────────────
    if followups:
        st.markdown('<hr class="msg-divider">', unsafe_allow_html=True)
        st.markdown('<div class="followup-label">Suggested follow-ups</div>', unsafe_allow_html=True)
        st.markdown('<div class="followup-btn">', unsafe_allow_html=True)
        fu_cols = st.columns(len(followups))
        for i, (col, fq) in enumerate(zip(fu_cols, followups)):
            with col:
                if st.button(fq, key=f"fu_{idx}_{i}", use_container_width=True):
                    st.session_state.pending_prompt = fq
                    st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)


# ── Session state ─────────────────────────────────────────────────────────────
for key, default in [
    ("dfs", {}), ("merged_df", None), ("join_cols", {}),
    ("chat_history", []), ("df_summary", ""), ("pending_prompt", None),
    # multi-level join state
    ("l2_joins", []),          # list of {left, right, on, result_name}
    ("available_dfs", {}),     # raw + any l2-merged datasets available for L1 join
]:
    if key not in st.session_state:
        st.session_state[key] = default

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="padding: 24px 20px 0 20px;">
        <div class="wordmark-bar">
            <div class="wordmark-icon">🔬</div>
            <div>
                <div class="wordmark-text">ExcelLens</div>
                <div class="wordmark-sub">Excel Intelligence</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    has_data   = bool(st.session_state.dfs)
    has_merged = st.session_state.merged_df is not None

    st.markdown('<div style="padding: 0 14px;"><div class="sidebar-label">Workflow</div>', unsafe_allow_html=True)
    st.markdown(f"""
        <div class="nav-item {'active' if not has_data else ''}"><span class="nav-icon">📁</span> Upload Data</div>
        <div class="nav-item {'active' if has_data and not has_merged else ''}"><span class="nav-icon">🔗</span> Join Datasets</div>
        <div class="nav-item {'active' if has_merged else ''}"><span class="nav-icon">💬</span> Query & Analyze</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div style="padding: 0 14px;"><div class="sidebar-label">Data Source</div></div>', unsafe_allow_html=True)

    uploaded_files = st.file_uploader("Drop Excel files here", type=["xlsx", "xls"],
                                      accept_multiple_files=True, label_visibility="collapsed")
    if uploaded_files:
        st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)
        if st.button("⟳  Load & Analyze", use_container_width=True):
            with st.spinner("Scanning files…"):
                st.session_state.dfs = load_multiple_excels(uploaded_files)
                st.session_state.join_cols = find_auto_join_columns(st.session_state.dfs)
            st.rerun()

    if has_data:
        st.markdown(f'<div style="padding:16px 14px 0 14px;"><div class="status-pill"><span class="dot"></span>{len(st.session_state.dfs)} datasets loaded</div></div>', unsafe_allow_html=True)
    if has_merged:
        r, c = st.session_state.merged_df.shape
        st.markdown(f'<div style="padding:6px 14px 0 14px;"><div class="status-pill"><span class="dot"></span>Master: {r:,} rows × {c} cols</div></div>', unsafe_allow_html=True)

    st.markdown("""
    <div style="position:fixed;bottom:0;left:0;width:260px;padding:16px 20px;
                border-top:1px solid #1e2d3d;background:#0d1117;">
        <div style="font-family:'IBM Plex Mono',monospace;font-size:0.6rem;
                    color:rgb(196 229 227);line-height:1.8;">
            LangGraph · Groq · Pandas · Streamlit<br>
            <span style="color:#1e2d3d;">────────────────</span><br>
            ExcelLens v1.0 · March 2026
        </div>
    </div>
    """, unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN CONTENT
# ═══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<div style="background:#080c14;border-bottom:1px solid #1e2d3d;padding:10px 40px;
            display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px;">
    <div style="display:flex;align-items:center;gap:24px;flex-wrap:wrap;">
        <div style="display:flex;align-items:center;gap:8px;">
            <span style="font-size:1rem;">🔬</span>
            <span style="font-family:'Syne',sans-serif;font-weight:800;font-size:0.9rem;color:#f8fafc;letter-spacing:-0.02em;">ExcelLens</span>
            <span style="font-family:'IBM Plex Mono',monospace;font-size:0.6rem;color:#334155;background:#111827;border:1px solid #1e2d3d;border-radius:4px;padding:1px 6px;">v1.0</span>
        </div>
        <div style="display:flex;gap:20px;align-items:center;">
            <div style="font-family:'IBM Plex Mono',monospace;font-size:0.67rem;color:rgb(186 198 215);"><span style="color:#334155;">◆</span>&nbsp;Upload <span style="color:#2563eb;">multiple Excel files</span></div>
            <div style="font-family:'IBM Plex Mono',monospace;font-size:0.67rem;color:rgb(186 198 215);"><span style="color:#334155;">◆</span>&nbsp;Auto-detect <span style="color:#2563eb;">shared columns</span> &amp; join</div>
            <div style="font-family:'IBM Plex Mono',monospace;font-size:0.67rem;color:rgb(186 198 215);"><span style="color:#334155;">◆</span>&nbsp;Ask <span style="color:#2563eb;">plain-English questions</span> via AI agent</div>
            <div style="font-family:'IBM Plex Mono',monospace;font-size:0.67rem;color:rgb(186 198 215);"><span style="color:#334155;">◆</span>&nbsp;Results as <span style="color:#2563eb;">dynamic tables</span> + Excel export</div>
        </div>
    </div>
    <div style="display:flex;gap:8px;align-items:center;">
        <span style="font-family:'IBM Plex Mono',monospace;font-size:0.6rem;color:#334155;letter-spacing:0.1em;">POWERED BY</span>
        <span style="font-family:'IBM Plex Mono',monospace;font-size:0.62rem;background:#0f1629;border:1px solid #1e2d3d;border-radius:4px;padding:2px 8px;color:#818cf8;">LangGraph</span>
        <span style="font-family:'IBM Plex Mono',monospace;font-size:0.62rem;background:#0f1629;border:1px solid #1e2d3d;border-radius:4px;padding:2px 8px;color:#818cf8;">Groq</span>
        <span style="font-family:'IBM Plex Mono',monospace;font-size:0.62rem;background:#0f1629;border:1px solid #1e2d3d;border-radius:4px;padding:2px 8px;color:#818cf8;">Pandas</span>
    </div>
</div>
""", unsafe_allow_html=True)

main_pad = "padding: 32px 40px 40px 40px;"

if not has_data:
    st.markdown(f'<div style="{main_pad}">', unsafe_allow_html=True)
    st.markdown("""
    <div style="max-width:640px;margin:80px auto;text-align:center;">
        <div style="font-family:'IBM Plex Mono',monospace;font-size:0.65rem;letter-spacing:0.2em;text-transform:uppercase;color:#2563eb;margin-bottom:14px;">Excel Intelligence Platform</div>
        <div style="font-family:'Syne',sans-serif;font-size:3rem;font-weight:800;color:#f8fafc;letter-spacing:-0.04em;line-height:1.05;margin-bottom:20px;">Turn spreadsheets<br>into answers.</div>
        <div style="font-family:'IBM Plex Sans',sans-serif;font-size:0.95rem;color:rgb(186 198 215);line-height:1.7;margin-bottom:40px;">
            Upload multiple Excel files. ExcelLens auto-detects shared columns, merges your data,
            and lets you ask plain-English statistical questions — powered by a live AI agent.
        </div>
        <div style="display:flex;gap:24px;justify-content:center;flex-wrap:wrap;">
            <div style="background:#0d1117;border:1px solid #1e2d3d;border-radius:10px;padding:20px 24px;text-align:left;flex:1;min-width:160px;">
                <div style="font-size:1.4rem;margin-bottom:10px;">📁</div>
                <div style="font-family:'Syne',sans-serif;font-weight:700;font-size:0.88rem;color:#e2e8f0;margin-bottom:6px;">Upload</div>
                <div style="font-size:0.75rem;color:rgb(186 198 215);font-family:'IBM Plex Sans',sans-serif;">Drop one or more .xlsx files into the sidebar</div>
            </div>
            <div style="background:#0d1117;border:1px solid #1e2d3d;border-radius:10px;padding:20px 24px;text-align:left;flex:1;min-width:160px;">
                <div style="font-size:1.4rem;margin-bottom:10px;">🔗</div>
                <div style="font-family:'Syne',sans-serif;font-weight:700;font-size:0.88rem;color:#e2e8f0;margin-bottom:6px;">Auto-Join</div>
                <div style="font-size:0.75rem;color:rgb(186 198 215);font-family:'IBM Plex Sans',sans-serif;">Shared columns detected and merged automatically</div>
            </div>
            <div style="background:#0d1117;border:1px solid #1e2d3d;border-radius:10px;padding:20px 24px;text-align:left;flex:1;min-width:160px;">
                <div style="font-size:1.4rem;margin-bottom:10px;">💬</div>
                <div style="font-family:'Syne',sans-serif;font-weight:700;font-size:0.88rem;color:#e2e8f0;margin-bottom:6px;">Ask</div>
                <div style="font-size:0.75rem;color:rgb(186 198 215);font-family:'IBM Plex Sans',sans-serif;">Ask statistical questions in plain English</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Multi-level Join Builder
# ══════════════════════════════════════════════════════════════════════════════
st.markdown(f'<div style="{main_pad}">', unsafe_allow_html=True)
st.markdown('<div class="section-eyebrow">Step 01</div><div class="section-title">Loaded Datasets</div>', unsafe_allow_html=True)

# ── Dataset preview ────────────────────────────────────────────────────────────
for i, (name, df) in enumerate(st.session_state.dfs.items()):
    with st.expander(f"📄  {name}   ·   {df.shape[0]:,} rows × {df.shape[1]} cols", expanded=(i == 0)):
        st.dataframe(df.head(5), width='stretch', hide_index=True)
        col_list = "  ".join([f"`{c}`" for c in df.columns.tolist()])
        st.markdown(f"<div style='margin-top:8px;font-family:IBM Plex Mono,monospace;font-size:0.68rem;color:#475569;'>{col_list}</div>", unsafe_allow_html=True)

st.markdown("<hr>", unsafe_allow_html=True)
st.markdown('<div class="section-eyebrow">Step 01b</div><div class="section-title">Join Configuration</div>', unsafe_allow_html=True)

# Keep available_dfs in sync with raw dfs + any completed L2 merges
all_available = dict(st.session_state.dfs)
for j in st.session_state.l2_joins:
    if j["result_name"] not in all_available:
        all_available[j["result_name"]] = j["result_df"]
st.session_state.available_dfs = all_available

dataset_names = list(st.session_state.dfs.keys())

# ── L2 Join Panel ──────────────────────────────────────────────────────────────
st.markdown("""
<div style="background:#080c14;border:1px solid #1e2d3d;border-radius:12px;padding:20px;margin-bottom:20px;">
    <div style="font-family:'IBM Plex Mono',monospace;font-size:0.62rem;letter-spacing:0.15em;
                text-transform:uppercase;color:#7c3aed;margin-bottom:4px;">Level 2 Pre-joins</div>
    <div style="font-family:'IBM Plex Sans',sans-serif;font-size:0.78rem;color: rgb(186, 198, 215);;margin-bottom:16px;">
        Join datasets that are not directly connected to your main table.
        Their merged result becomes available for the Level 1 join below.
    </div>
""", unsafe_allow_html=True)

# Show existing L2 joins as a live pipeline
if st.session_state.l2_joins:
    for idx, j in enumerate(st.session_state.l2_joins):
        c1, c2, c3, c4, c5 = st.columns([2, 0.4, 2, 2, 0.8])
        with c1:
            st.markdown(f"""<div style="background:#0d1117;border:1px solid #2d3a5c;border-radius:8px;
                padding:8px 12px;font-family:IBM Plex Mono,monospace;font-size:0.72rem;color:#818cf8;">
                📄 {j["left"]}</div>""", unsafe_allow_html=True)
        with c2:
            st.markdown('<div style="text-align:center;padding-top:10px;color:#475569;font-size:1rem;">⟷</div>', unsafe_allow_html=True)
        with c3:
            st.markdown(f"""<div style="background:#0d1117;border:1px solid #2d3a5c;border-radius:8px;
                padding:8px 12px;font-family:IBM Plex Mono,monospace;font-size:0.72rem;color:#818cf8;">
                📄 {j["right"]}</div>""", unsafe_allow_html=True)
        with c4:
            st.markdown(f"""<div style="background:#0d1117;border:1px solid #059669;border-radius:8px;
                padding:8px 12px;font-family:IBM Plex Mono,monospace;font-size:0.68rem;color:#34d399;">
                ✓ {j["result_name"]}  <span style="color:#334155;">on: {j["on"]}</span></div>""", unsafe_allow_html=True)
        with c5:
            if st.button("✕", key=f"del_l2_{idx}", help="Remove this L2 join"):
                st.session_state.l2_joins.pop(idx)
                st.rerun()
    st.markdown('<div style="height:12px;"></div>', unsafe_allow_html=True)

# Add new L2 join form
with st.expander("➕  Add Level 2 Pre-join", expanded=(len(st.session_state.l2_joins) == 0)):
    lc1, lc2, lc3 = st.columns([2, 2, 2])
    with lc1:
        l2_left = st.selectbox("Left dataset", options=dataset_names, key="l2_left")
    with lc2:
        l2_right_opts = [n for n in dataset_names if n != l2_left]
        l2_right = st.selectbox("Right dataset", options=l2_right_opts, key="l2_right") if l2_right_opts else None
    with lc3:
        if l2_right:
            common = find_common_cols(
                st.session_state.dfs[l2_left],
                st.session_state.dfs[l2_right]
            )
            if common:
                l2_on = st.selectbox("Join on column", options=common, key="l2_on")
            else:
                st.warning("No common columns found between these two datasets.")
                l2_on = None
        else:
            l2_on = None

    if l2_right and l2_on:
        result_default = f"{l2_left[:8]}_{l2_right[:8]}_merged"
        l2_result_name = st.text_input("Result name", value=result_default, key="l2_result_name")
        if st.button("🔗  Create L2 Pre-join", use_container_width=True, key="add_l2"):
            with st.spinner(f"Joining {l2_left} ⟷ {l2_right}…"):
                result_df = do_join(
                    st.session_state.dfs[l2_left],
                    st.session_state.dfs[l2_right],
                    l2_on
                )
            st.session_state.l2_joins.append({
                "left":        l2_left,
                "right":       l2_right,
                "on":          l2_on,
                "result_name": l2_result_name,
                "result_df":   result_df,
            })
            st.session_state.available_dfs[l2_result_name] = result_df
            st.success(f"✅ Created **{l2_result_name}** — {result_df.shape[0]:,} rows × {result_df.shape[1]} cols")
            st.rerun()

st.markdown('</div>', unsafe_allow_html=True)

# ── L1 Join Panel ──────────────────────────────────────────────────────────────
st.markdown("""
<div style="background:#080c14;border:1px solid #1e2d3d;border-radius:12px;padding:20px;">
    <div style="font-family:'IBM Plex Mono',monospace;font-size:0.62rem;letter-spacing:0.15em;
                text-transform:uppercase;color:#2563eb;margin-bottom:4px;">Level 1 — Main Join</div>
    <div style="font-family:'IBM Plex Sans',sans-serif;font-size:0.78rem;color: rgb(186, 198, 215);margin-bottom:16px;">
        Select your primary table and join it with any dataset (including L2 pre-joined results above).
    </div>
""", unsafe_allow_html=True)

available_names = list(st.session_state.available_dfs.keys())

lc1, lc2 = st.columns([1, 2])
with lc1:
    main_df_name = st.selectbox("Primary (base) table", options=available_names, key="main_select")
with lc2:
    other_names = [n for n in available_names if n != main_df_name]
    join_targets = st.multiselect(
        "Join with (select one or more)",
        options=other_names,
        default=other_names,
        key="l1_targets",
    )

# Show auto-detected join keys for selected targets
if join_targets:
    all_join_dfs = {n: st.session_state.available_dfs[n] for n in [main_df_name] + join_targets}
    detected_cols = find_auto_join_columns(all_join_dfs)
    if detected_cols:
        badges = "".join([f'<span class="join-badge">⟷ {col}</span>' for col in list(detected_cols.keys())[:8]])
        st.markdown(f'<div style="margin:10px 0 4px 0;font-size:0.7rem;color: rgb(186, 198, 215);font-family:IBM Plex Mono,monospace;">Auto-detected join keys:</div>{badges}', unsafe_allow_html=True)
    else:
        st.warning("No common columns found between selected datasets.")

st.markdown('<div style="height:12px;"></div>', unsafe_allow_html=True)

if st.button("🚀  Build Master Dataset", use_container_width=True, key="build_master"):
    with st.spinner("Building master dataset…"):
        main_df     = st.session_state.available_dfs[main_df_name]
        other_dfs   = {n: st.session_state.available_dfs[n] for n in join_targets}
        all_join_dfs = {main_df_name: main_df, **other_dfs}
        common_cols = list(find_auto_join_columns(all_join_dfs).keys())
        st.session_state.merged_df  = smart_join_dfs(main_df, other_dfs, common_cols)
        st.session_state.df_summary = get_df_summary(st.session_state.merged_df)
    st.rerun()

st.markdown('</div>', unsafe_allow_html=True)

# ── Step 2: Master dataset ─────────────────────────────────────────────────────
if st.session_state.merged_df is not None:
    mdf = st.session_state.merged_df
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="section-eyebrow">Step 02</div><div class="section-title">Master Dataset</div>', unsafe_allow_html=True)

    size_mb = mdf.memory_usage(deep=True).sum() / 1e6
    st.markdown(f"""
    <div class="metric-row">
        <div class="metric-card"><div class="metric-label">Total Rows</div><div class="metric-value">{mdf.shape[0]:,}</div></div>
        <div class="metric-card accent"><div class="metric-label">Columns</div><div class="metric-value">{mdf.shape[1]}</div></div>
        <div class="metric-card green"><div class="metric-label">Memory</div><div class="metric-value">{size_mb:.1f}<span style="font-size:1rem;color:#475569"> MB</span></div></div>
        <div class="metric-card amber"><div class="metric-label">Join Keys</div><div class="metric-value">{len(st.session_state.join_cols)}</div></div>
    </div>
    """, unsafe_allow_html=True)

    st.dataframe(mdf.head(10), width='stretch', hide_index=True)
    tags = "".join([f'<span class="col-tag">{c}</span>' for c in mdf.columns.tolist()])
    st.markdown(f'<div style="margin-top:12px;">{tags}</div>', unsafe_allow_html=True)

# ── Step 3: Chat ───────────────────────────────────────────────────────────────
if st.session_state.merged_df is not None:
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="section-eyebrow">Step 03</div><div class="section-title">Query Your Data</div>', unsafe_allow_html=True)

    st.markdown("""
    <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:16px;">
        <span style="font-family:'IBM Plex Mono',monospace;font-size:0.62rem;color:#475569;letter-spacing:0.1em;text-transform:uppercase;align-self:center;">Try:</span>
        <span class="join-badge">Top 5 by salary</span>
        <span class="join-badge">Average score by dept</span>
        <span class="join-badge">Count employees aged 35+</span>
        <span class="join-badge">Total budget by region</span>
    </div>
    """, unsafe_allow_html=True)

    # ── Chat history ──────────────────────────────────────────────────────────
    chat_container = st.container(height=520)
    with chat_container:
        if not st.session_state.chat_history:
            st.markdown("""
            <div class="empty-state">
                <div class="empty-icon">💬</div>
                <div class="empty-title">No queries yet</div>
                <div class="empty-sub">Ask a statistical question about your data below</div>
            </div>
            """, unsafe_allow_html=True)
        else:
            history = st.session_state.chat_history
            for idx, chat in enumerate(history[-16:]):
                role = chat["role"]
                actual_idx = len(history) - len(history[-16:]) + idx

                with st.chat_message(role, avatar="👤" if role == "user" else "🔬"):
                    if role == "user":
                        st.markdown(f'<div class="user-msg">{chat["content"]}</div>', unsafe_allow_html=True)
                    else:
                        is_latest = (actual_idx == len(history) - 1)
                        render_answer(chat, actual_idx, is_latest)

    # ── Process pending prompt (from follow-up button click) ──────────────────
    if st.session_state.pending_prompt:
        prompt = st.session_state.pending_prompt
        st.session_state.pending_prompt = None
    else:
        prompt = st.chat_input("e.g. 'Show top 10 employees by performance score with their department'")

    if prompt:
        st.session_state.chat_history.append({"role": "user", "content": prompt})

        with st.chat_message("user", avatar="👤"):
            st.markdown(f'<div class="user-msg">{prompt}</div>', unsafe_allow_html=True)

        with st.chat_message("assistant", avatar="🔬"):
            with st.spinner("Running analysis…"):
                structured = analyze_query_structured(
                    st.session_state.merged_df,
                    prompt,
                    st.session_state.df_summary,
                )
            assistant_entry = {
                "role": "assistant",
                "content": structured.get("summary", ""),
                "question": prompt,
                "structured": structured,
            }
            st.session_state.chat_history.append(assistant_entry)
            render_answer(assistant_entry, len(st.session_state.chat_history) - 1, is_latest=True)
            st.rerun()

st.markdown('</div>', unsafe_allow_html=True)